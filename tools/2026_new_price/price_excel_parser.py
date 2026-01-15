import sys
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl

KAZEI_KBN = {
    "0": "課税",
    "1": "非課税",
}

WARIMASHI_KBN = {
    "1": "基本料金",
    "23": "BS託児　法人会員　基本料金",
    "2": "早朝割増",
    "3": "夜間割増",
    "4": "深夜割増",
    "5": "デイケア延長",
    "6": "デイケア日数追加",
    "7": "教室日数追加",
    "8": "ﾊﾞｯｸｱｯﾌﾟ",
    "9": "四季",
    "10": "宴会イベント",
    "11": "託児　兄弟追加時間料金",
    "12": "予約手数料　２日前",
    "13": "予約手数料　１日前",
    "14": "予約手数料　当日",
    "15": "予約変更料　１日前",
    "16": "予約変更料　当日",
    "17": "託児ｷｬﾝｾﾙ　１日前",
    "18": "託児ｷｬﾝｾﾙ　当日",
    "19": "託児　出張費　23:00～翌7:00",
    "24": "託児　出張費　21:00～翌8:00",
    "25": "託児　出張費　22:00～翌8:00",
    "26": "託児　出張費　22:00～翌9:00",
    "20": "託児　出張費　7:00～8:00，22:00～23:00",
    "21": "クレアナーサリー延長",
    "22": "居宅訪問延長",
    "0": "その他",
}


UNIT_KBN = {
    "0": "",
    "1": "年",
    "2": "月",
    "3": "日",
    "4": "時間",
    "5": "分",
    "6": "家族",
    "7": "契約",
    "8": "初回",
    "9": "人",
    "10": "回",
    "11": "枚",
    "12": "往復",
    "13": "10分",
    "14": "30分",
    "15": "3時間",
}

HINMOKU_KBN = {
    "0": "なし",
    "1": "基本",
    "2": "デイケア",
    "3": "バックアップ",
    "4": "兄弟",
    "5": "指名",
    "6": "チューター、ドゥラー、ハウスキーピング",
    "7": "入浴",
    "8": "出張",
    "9": "予約",
    "10": "食事",
    "11": "雑費",
}

HINMOKU2_KBN = {
    "0": "なし",
    "1": "ハウスキーピング(HK)",
    "2": "チューター、ドゥラー、外国語対応(TDB)",
    "3": "指名",
    "4": "入浴",
}


def copy_template_with_timestamp() -> Path:
    """Copy the template.xlsx to output directory with a timestamped filename."""
    repo_root = Path("/workspaces/canals-reader-api")
    template_path = repo_root / "assets" / "template.xlsx"
    output_dir = repo_root / "tools" / "2026_new_price" / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    dest_path = output_dir / f"{ts}.xlsx"

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    shutil.copy2(str(template_path), str(dest_path))
    return dest_path


def read_price_excel(file_path: str, header_row: int = 1, sheet_name: Optional[str] = None) -> List[Dict]:
    """
    Read an Excel file and convert rows to a list of dictionaries using the header row as keys.

    Args:
        file_path: path to the .xlsx file
        header_row: 1-based index of the header row (defaults to 1)
        sheet_name: optional sheet name to read; if None the active sheet is used

    Returns:
        List of dicts, where each dict corresponds to a row mapping header->cell value.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # openpyxl is 1-based for rows/columns
    header_row_idx = header_row
    headers = []
    for cell in ws[header_row_idx]:
        # Normalize header: strip whitespace and convert empty to None
        val = cell.value
        if isinstance(val, str):
            val = val.strip()
        headers.append(val)

    records: List[Dict] = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # stop if entire row is empty
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        rec: Dict = {}
        for idx, val in enumerate(row):
            # If header is missing or None, skip that column
            if idx >= len(headers):
                # extra column without header; use numeric key
                key = f"col_{idx + 1}"
            else:
                key = headers[idx]
                if key is None or (isinstance(key, str) and key == ""):
                    key = f"col_{idx + 1}"
            rec[key] = val
        records.append(rec)

    return records


def parse_for_result(record: Dict) -> Dict:
    """Transform the input records for the result Excel."""
#     data = {
#     "ＩＤ": "3685",
#     "拠点": "36",
#     "拠点名": "家事代行",
#     "品目名": "当社ギフト券",
#     "品目名（英語）": null,
#     "対象時間From": null,
#     "対象時間To": null,
#     "料金（税抜）": "-1000",
#     "変更後": null,
#     "開始日": null,
#     "割増区分": "0",
#     "自動計算で使用する": "0",
#     "対象時間区分 平日": "1",
#     "対象時間区分 土": "1",
#     "対象時間区分 日・祝": "1",
#     "最小受注": null,
#     "単位": "0",
#     "単位区分": "0",
#     "単位（テキスト入力）": null,
#     "課税区分 (0:課税 1:非課税)": "1",
#     "税率": "0",
#     "ポイント率": "1",
#     "有効期限From": null,
#     "有効期限To": null,
#     "備考": null,
#     "優待フラグ": "0",
#     "勘定科目コード": "888",
#     "勘定科目名": "優待割引",
#     "請求方法 法人": "0",
#     "請求方法 個人": "0",
#     "使用区分 入会金": "0",
#     "使用区分 年会費": "0",
#     "使用区分 受注・前受 デイケア": "0",
#     "使用区分 受注・前受 教室": "0",
#     "使用区分 受注・前受 AKC": "0",
#     "使用区分 受注・前受 バックアップ": "0",
#     "使用区分 初回請求分": "0",
#     "使用区分 予約 デイケア": "0",
#     "使用区分 予約 デイケア日数追加": "0",
#     "使用区分 予約 教室日数追加": "0",
#     "使用区分 予約 教室": "0",
#     "使用区分 予約 一時預かり/ホテル託児": "0",
#     "使用区分 予約 AKC": "0",
#     "使用区分 予約 バックアップ": "0",
#     "使用区分 予約 劇団四季": "0",
#     "使用区分 予約 宴会イベント": "0",
#     "使用区分 予約 家事代行": "0",
#     "使用区分 実施 デイケア": "0",
#     "使用区分 実施 デイケア日数追加": "0",
#     "使用区分 実施 教室日数追加": "0",
#     "使用区分 実施 教室": "0",
#     "使用区分 実施 一時預かり/ホテル託児": "0",
#     "使用区分 実施 AKC": "0",
#     "使用区分 実施 バックアップ": "0",
#     "使用区分 実施 劇団四季": "0",
#     "使用区分 実施 宴会イベント": "0",
#     "使用区分 実施 家事代行": "0",
#     "使用区分 割引チケット": "1",
#     "顧客区分 プレミア": "1",
#     "顧客区分 ビジター": "1",
#     "顧客区分 法人メンバー": "1",
#     "顧客区分 支店会員": "1",
#     "顧客区分 法人": "0",
#     "顧客区分 BS会員": "0",
#     "顧客区分 三越伊勢丹グループ各種カード会員": "0",
#     "補助科目コード": "0",
#     "Plan-B 超過基準時間": "0",
#     "割引率": "0",
#     "割引率フラグ": "0",
#     "法人番号": "0",
#     "キャンセルフィーに含む 前日": "0",
#     "キャンセルフィーに含む 当日": "0",
#     "表示優先順位": "0",
#     "定率割引・割増に含む": "0",
#     "品目区分 (0:なし 1:基本 2:デイケア 3:バックアップ 4:兄弟 5:指名 6:チューター、ドゥラー、ハウスキーピング 7:入浴 8:出張 9:予約 10:食事 11:雑費)": "0",
#     "col_76": null
#   }
    # 品目名
    item_name = f'{record.get("品目名")}【{record.get("開始日")}】'

    return {

        "拠点": record.get("拠点名"),
        "品目名": item_name,
        "料金": record.get("変更後"),
        "課税区分": KAZEI_KBN.get(str(record.get("課税区分 (0:課税 1:非課税)")), "課税"),
        "対象時間From": record.get("対象時間From"),
        "対象時間To": record.get("対象時間To"),
        "割増区分": WARIMASHI_KBN.get(str(record.get("割増区分")), "その他"),
        "自動計算フラグ": "0",
        "対象時間区分": build_taisyo_jikan_kbn(record),
        "最小受注": record.get("最小受注"),
        "単位数値": record.get("単位"),
        "単位": UNIT_KBN.get(str(record.get("単位区分")), ""),
        "単位テキスト": record.get("単位（テキスト入力）"),
        "ポイント率": record.get("ポイント率"),
        "税率": "10",
        "備考": record.get("備考"),
        "優待フラグ": record.get("優待フラグ"),
        "勘定科目コード": record.get("勘定科目コード"),
        "勘定科目名": record.get("勘定科目名"),
        "補助科目コード": record.get("補助科目コード"),
        "請求方法": build_seikyu_houhou(record),
        "PlanB超過基準時間": record.get("Plan-B 超過基準時間"),
        "割引率": record.get("割引率"),
        "割引フラグ": record.get("割引率フラグ"),
        "法人番号": record.get("法人番号"),
        "キャンセルフィー": build_cancel_fee(record),
        "表示優先度": record.get("表示優先順位"),
        "定率割引フラグ": record.get("定率割引・割増に含む"),
        "品目区分": HINMOKU_KBN.get(str(record.get("品目区分")), "なし"),
        "品目区分2": "なし",
        "使用区分": "",
        "入会金": record.get("使用区分 入会金"),
        "年会費": record.get("使用区分 年会費"),
        "受注・前受 デイケア": record.get("使用区分 受注・前受 デイケア"),
        "受注・前受 教室": record.get("使用区分 受注・前受 教室"),
        "受注・前受 AKC": record.get("使用区分 受注・前受 AKC"),
        "受注・前受 バックアップ": record.get("使用区分 受注・前受 バックアップ"),
        "初回請求分": record.get("使用区分 初回請求分"),
        "予約 デイケア": record.get("使用区分 予約 デイケア"),
        "予約 デイケア日数追加": record.get("使用区分 予約 デイケア日数追加"),
        "予約 教室日数追加": record.get("使用区分 予約 教室日数追加"),
        "予約 教室": record.get("使用区分 予約 教室"),
        "予約 一時預かり/ホテル託児": record.get("使用区分 予約 一時預かり/ホテル託児"),
        "予約 AKC": record.get("使用区分 予約 AKC"),
        "予約 バックアップ": record.get("使用区分 予約 バックアップ"),
        "予約 劇団四季": record.get("使用区分 予約 劇団四季"),
        "予約 宴会イベント": record.get("使用区分 予約 宴会イベント"),
        "予約 家事代行": record.get("使用区分 予約 家事代行"),
        "実施 デイケア": record.get("使用区分 実施 デイケア"),
        "実施 デイケア日数追加": record.get("使用区分 実施 デイケア日数追加"),
        "実施 教室日数追加": record.get("使用区分 実施 教室日数追加"),
        "実施 教室": record.get("使用区分 実施 教室"),
        "実施 一時預かり/ホテル託児": record.get("使用区分 実施 一時預かり/ホテル託児"),
        "実施 AKC": record.get("使用区分 実施 AKC"),
        "実施 バックアップ": record.get("使用区分 実施 バックアップ"),
        "実施 劇団四季": record.get("使用区分 実施 劇団四季"),
        "実施 宴会イベント": record.get("使用区分 実施 宴会イベント"),
        "実施 家事代行": record.get("使用区分 実施 家事代行"),
        "割引チケット": record.get("使用区分 割引チケット"),
        "顧客区分": "",
        "プレミア": record.get("顧客区分 プレミア"),
        "ビジター": record.get("顧客区分 ビジター"),
        "法人メンバー": record.get("顧客区分 法人メンバー"),
        "支店会員": record.get("顧客区分 支店会員"),
        "法人": record.get("顧客区分 法人"),
        "BS会員": record.get("顧客区分 BS会員"),
        "優待会員": record.get("顧客区分 三越伊勢丹グループ各種カード会員"),
    }


def build_taisyo_jikan_kbn(record: Dict) -> str:
    kbns = []
    if record.get("対象時間区分 平日") == "1":
        kbns.append("平日")
    if record.get("対象時間区分 土") == "1":
        kbns.append("土")
    if record.get("対象時間区分 日・祝") == "1":
        kbns.append("日・祝")
    return ",".join(kbns)


def build_seikyu_houhou(record: Dict) -> str:
    kbns = []
    if record.get("請求方法 法人") == "1":
        kbns.append("法人")
    if record.get("請求方法 個人") == "1":
        kbns.append("個人")
    return ",".join(kbns)


def build_cancel_fee(record: Dict) -> str:
    kbns = []
    if record.get("キャンセルフィーに含む 前日") == "1":
        kbns.append("前日")
    if record.get("キャンセルフィーに含む 当日") == "1":
        kbns.append("当日")
    return ",".join(kbns)


def write_price_excel(file_path: str, records: List[Dict]) -> int:
    # --- write records into the copied template (Sheet1) starting at row 2 ---
    try:
        wb_out = openpyxl.load_workbook(file_path)
        # prefer explicit 'Sheet1' if present, else active
        if "Sheet1" in wb_out.sheetnames:
            ws_out = wb_out["Sheet1"]
        else:
            ws_out = wb_out.active

        # read header names from row 1
        headers = [cell.value for cell in ws_out[1]]
        # normalize headers: fallback col_N for empty headers
        norm_headers = []
        for idx, h in enumerate(headers):
            if h is None or (isinstance(h, str) and h.strip() == ""):
                norm_headers.append(f"col_{idx + 1}")
            else:
                norm_headers.append(h)

        # write each record starting from row 2
        start_row = 2
        for r_idx, rec in enumerate(records, start=start_row):
            for c_idx, key in enumerate(norm_headers, start=1):
                val = rec.get(key)
                # write value directly (openpyxl will handle numbers/dates/strings)
                ws_out.cell(row=r_idx, column=c_idx, value=val)

        wb_out.save(file_path)
        print(
            f"Wrote {len(records)} rows to: {file_path}", file=sys.stderr)
    except Exception as e:
        print(
            f"Error writing to output Excel {file_path}: {e}", file=sys.stderr)
        return 3


def main(argv: Optional[List[str]] = None) -> int:
    """CLI entrypoint. Prints JSON array to stdout."""
    if argv is None:
        argv = sys.argv
    if len(argv) >= 2:
        path = argv[1]
    else:
        # default source path used in this repo
        path = "/workspaces/canals-reader-api/tools/2026_new_price/source/2026年料金_手修正.xlsx"
    # --- copy template to output with timestamped filename ---
    dest_path = copy_template_with_timestamp()
    print(f"Copied template to: {dest_path}", file=sys.stderr)

    try:
        read_records = read_price_excel(path)
        # Filter to only records with "変更後" value
        read_records = [rec for rec in read_records if rec.get("変更後")]
        # Print JSON with ensure_ascii=False so Japanese characters are preserved
        # print(json.dumps(read_records, ensure_ascii=False, default=str, indent=2))

        write_records = [parse_for_result(rec) for rec in read_records]

        res = write_price_excel(str(dest_path), write_records)
        if res != 0:
            return res

        return 0

    except Exception as e:
        print(f"Error reading Excel: {e}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
