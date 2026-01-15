# from dateutil.relativedelta import relativedelta
import sys
import openpyxl
from typing import List, Dict, Optional

# from logics.common.const import DATE_FORMAT
from logics.utils.selenium_util import SeleniumUtil

# from logics.pages.sample_page import SamplePage
from logics.pages.canals_login_page import CanalsLoginPage
from logics.pages.canals_menu_page import CanalsMenuPage
from logics.pages.canals_ryoukin_edit_page import CanalsRyokinEditPage
from logics.pages.canals_ryoukin_list_page import CanalsRyokinListPage

HINMOKU_KBN = {
    'なし': '0',
    '基本': '1',
    'デイケア': '2',
    'バックアップ': '3',
    '兄弟': '4',
    '指名': '5',
    'チューター、ドゥラー、ハウスキーピング': '6',
    '入浴': '7',
    '出張': '8',
    '予約': '9',
    '食事': '10',
    '雑費': '11',
}

HINMOKU2_KBN = {
    'なし': '0',
    'ハウスキーピング(HK)': '1',
    'チューター、ドゥラー、外国語対応(TDB)': '2',
    '指名': '3',
    '入浴': '4',
}


def read_pages(env, file_path):

    records = read_price_excel(file_path, header_row=1)
    print(f"Read {len(records)} records from Excel")

    selenium = None
    try:
        # - ドライバーの初期化
        selenium = SeleniumUtil()
        selenium.build_pc_driver()

        config = make_config(env)
        pages = init_pages(selenium, config)

        # - ログイン処理
        _login(pages, config)

        # レコード登録処理
        for idx, record in enumerate(records):
            print(f"start record: [{idx + 2}] {record['品目名']}")
            _charge_register_one_rec(record, pages)
            print(f"end record: [{idx + 2}] {record['品目名']}")

    finally:
        if selenium:
            selenium.dispose_driver()

    result = {
        'statusCode': 200,
        'body': 0
    }

    return result


def read_price_excel(file_path: str, header_row: int = 1, sheet_name: Optional[str] = None) -> List[Dict]:
    """
    Read an Excel file and convert rows to a list of dictionaries using the header row as keys.

    Args:
        file_path: path to the .xlsx file
        header_row: 1-based index of the header row (defaults to 1)
        sheet_name: optional sheet name to read; if None the active sheet is used

    Returns:
        List of dicts, where each dict corresponds to a row mapping header->cell value.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # openpyxl is 1-based for rows/columns
    header_row_idx = header_row
    headers = []
    for cell in ws[header_row_idx]:
        # Normalize header: strip whitespace and convert empty to None
        val = cell.value
        if isinstance(val, str):
            val = val.strip()
        headers.append(val)

    records: List[Dict] = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # stop if entire row is empty
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        rec: Dict = {}
        for idx, val in enumerate(row):
            # If header is missing or None, skip that column
            if idx >= len(headers):
                # extra column without header; use numeric key
                key = f"col_{idx + 1}"
            else:
                key = headers[idx]
                if key is None or (isinstance(key, str) and key == ""):
                    key = f"col_{idx + 1}"
            rec[key] = val
        records.append(rec)

    return records


def make_config(env):
    if env == 'prod':
        config = {
            "CANALS_DOMAIN": "cscs.alpha-co.com",
            "CANALS_LOGIN_ID": "5057",
            "CANALS_PASSWORD": "alpha5057",
        }
    else:
        config = {
            "CANALS_DOMAIN": "alpha.awstest.canals.jp",
            "CANALS_LOGIN_ID": "5057",
            "CANALS_PASSWORD": "alpha5057",
        }

    return config


def init_pages(selenium, config):
    pages = {
        "login_page": CanalsLoginPage(selenium, config),
        "menu_page": CanalsMenuPage(selenium, config),
        "ryoukin_edit_page": CanalsRyokinEditPage(selenium, config),
        "ryoukin_list_page": CanalsRyokinListPage(selenium, config),
    }

    return pages


def _login(pages, config):
    login_page = pages["login_page"]
    print("Login start...")
    # ログイン画面を開く
    login_page.open()
    login_page.login_button().wait_to_be_clickable()

    # ID/PASSの入力
    login_page.login_id().input(config["CANALS_LOGIN_ID"])
    login_page.password().input(config["CANALS_PASSWORD"])
    login_page.login_button().click()

    # メニュー画面が表示されるまで待機
    menu_page = pages["menu_page"]
    menu_page.logout_button().wait_to_be_clickable()

    print("Login successful")


def _charge_register_one_rec(record, pages):
    ryoukin_list_page = pages["ryoukin_list_page"]

    # 料金一覧画面を開く
    ryoukin_list_page.open()
    ryoukin_list_page.new_registration_button().wait_to_be_clickable()

    ryoukin_edit_page = pages["ryoukin_edit_page"]
    # 料金編集画面を開く
    ryoukin_list_page.new_registration_button().click()
    ryoukin_edit_page.register_button().wait_to_be_clickable()

    # データ入力
    _input_one_rec(record, ryoukin_edit_page)

    # 登録ボタン押下
    ryoukin_edit_page.register_button().click()
    ryoukin_edit_page.submit_button().wait_to_be_clickable()

    # 送信ボタン押下
    ryoukin_edit_page.submit_button().click()
    ryoukin_edit_page.back_button().wait_to_be_clickable()

    # item_name = _get_value(record, '品目名', '')
    # ryoukin_edit_page.screenshot("", item_name)


def _input_one_rec(record, ryoukin_edit_page):
    ryoukin_edit_page.center_select_box().select_text(_get_value(record, '拠点', ''))
    ryoukin_edit_page.item_name_input().input(_get_value(record, '品目名', ''))
    ryoukin_edit_page.charge_input().clear()
    ryoukin_edit_page.charge_input().input(int(_get_value(record, '料金', 0)))

    kazei_kbn = "0" if _get_value(record, '課税区分', '') == "課税" else "1"
    ryoukin_edit_page.taxable_radio(kazei_kbn).click()

    ryoukin_edit_page.time_from_input().input(_get_value(record, '対象時間From', ''))
    ryoukin_edit_page.time_to_input().input(_get_value(record, '対象時間To', ''))
    ryoukin_edit_page.warimashi_select_box().select_text(
        _get_value(record, '割増区分', 'その他'))

    if str(_get_value(record, '自動計算フラグ', '')) == '1':
        ryoukin_edit_page.auto_calc_checkbox().click()

    time_kbns = _get_value(record, '対象時間区分', '').split(',')
    if '平日' in time_kbns:
        ryoukin_edit_page.time_kbn_checkbox_weekday().click()
    if '土' in time_kbns:
        ryoukin_edit_page.time_kbn_checkbox_saturday().click()
    if '日・祝' in time_kbns:
        ryoukin_edit_page.time_kbn_checkbox_sunday_holiday().click()

    ryoukin_edit_page.min_order_input().input(_get_value(record, '最小受注', ''))
    ryoukin_edit_page.unit_value_input().clear()
    ryoukin_edit_page.unit_value_input().input(_get_value(record, '単位数値', ''))

    unit_kbn = _get_value(record, '単位', '')
    if not unit_kbn:
        unit_kbn = '選択'
    ryoukin_edit_page.unit_select_box().select_text(unit_kbn)
    ryoukin_edit_page.unit_text_input().input(_get_value(record, '単位テキスト', ''))

    ryoukin_edit_page.point_rate_input().clear()
    ryoukin_edit_page.point_rate_input().input(_get_value(record, 'ポイント率', ''))
    ryoukin_edit_page.remarks_input().input(_get_value(record, '備考', ''))

    if str(_get_value(record, '優待フラグ', '')) == '1':
        ryoukin_edit_page.special_flag_checkbox().click()

    ryoukin_edit_page.account_code_input().clear()
    ryoukin_edit_page.account_code_input().input(_get_value(record, '勘定科目コード', ''))
    ryoukin_edit_page.account_name_input().input(_get_value(record, '勘定科目名', ''))
    ryoukin_edit_page.sub_account_code_input().clear()
    ryoukin_edit_page.sub_account_code_input().input(
        _get_value(record, '補助科目コード', ''))

    bill_kbns = _get_value(record, '請求方法', '').split(',')
    if '法人' in bill_kbns:
        ryoukin_edit_page.billing_method_checkbox_corporate().click()
    if '個人' in bill_kbns:
        ryoukin_edit_page.billing_method_checkbox_individual().click()

    ryoukin_edit_page.planb_excess_standard_time_input().input(
        _get_value(record, 'PlanB超過基準時間', ''))

    ryoukin_edit_page.discount_rate_input().clear()
    ryoukin_edit_page.discount_rate_input().input(_get_value(record, '割引率', ''))

    if str(_get_value(record, '割引フラグ', '')) == '1':
        ryoukin_edit_page.discount_flag_checkbox().click()

    ryoukin_edit_page.corporate_number_input().clear()
    ryoukin_edit_page.corporate_number_input().input(_get_value(record, '法人番号', ''))

    cancel_fees = _get_value(record, 'キャンセルフィー', '').split(',')
    if '前日' in cancel_fees:
        ryoukin_edit_page.cancel_fee_checkbox_before_day().click()
    if '当日' in cancel_fees:
        ryoukin_edit_page.cancel_fee_checkbox_on_day().click()

    ryoukin_edit_page.display_priority_input().clear()
    ryoukin_edit_page.display_priority_input().input(_get_value(record, '表示優先度', ''))

    if str(_get_value(record, '定率割引フラグ', '')) == '1':
        ryoukin_edit_page.flat_rate_discount_flag_checkbox().click()

    hinmoku_kbn = HINMOKU_KBN.get(_get_value(record, '品目区分', ''), '0')
    ryoukin_edit_page.item_kbn_radio(hinmoku_kbn).click()
    hinmoku_kbn2 = HINMOKU2_KBN.get(_get_value(record, '品目区分2', ''), '0')
    ryoukin_edit_page.item_kbn2_radio(hinmoku_kbn2).click()

    # 使用区分
    if str(_get_value(record, '入会金', '')) == '1':
        ryoukin_edit_page.entrance_fee_checkbox().click()
    if str(_get_value(record, '年会費', '')) == '1':
        ryoukin_edit_page.annual_fee_checkbox().click()
    if str(_get_value(record, '受注・前受 デイケア', '')) == '1':
        ryoukin_edit_page.order_prepay_daycare_checkbox().click()
    if str(_get_value(record, '受注・前受 教室', '')) == '1':
        ryoukin_edit_page.order_prepay_classroom_checkbox().click()
    if str(_get_value(record, '受注・前受 AKC', '')) == '1':
        ryoukin_edit_page.order_prepay_akc_checkbox().click()
    if str(_get_value(record, '受注・前受 バックアップ', '')) == '1':
        ryoukin_edit_page.order_prepay_backup_checkbox().click()
    if str(_get_value(record, '初回請求分', '')) == '1':
        ryoukin_edit_page.first_billing_checkbox().click()
    if str(_get_value(record, '予約 デイケア', '')) == '1':
        ryoukin_edit_page.reservation_daycare_checkbox().click()
    if str(_get_value(record, '予約 デイケア日数追加', '')) == '1':
        ryoukin_edit_page.reservation_daycare_days_add_checkbox().click()
    if str(_get_value(record, '予約 教室日数追加', '')) == '1':
        ryoukin_edit_page.reservation_classroom_days_add_checkbox().click()
    if str(_get_value(record, '予約 教室', '')) == '1':
        ryoukin_edit_page.reservation_classroom_checkbox().click()
    if str(_get_value(record, '予約 一時預かり/ホテル託児', '')) == '1':
        ryoukin_edit_page.reservation_temporary_care_checkbox().click()
    if str(_get_value(record, '予約 AKC', '')) == '1':
        ryoukin_edit_page.reservation_akc_checkbox().click()
    if str(_get_value(record, '予約 バックアップ', '')) == '1':
        ryoukin_edit_page.reservation_backup_checkbox().click()
    if str(_get_value(record, '予約 劇団四季', '')) == '1':
        ryoukin_edit_page.reservation_gekidan_shiki_checkbox().click()
    if str(_get_value(record, '予約 宴会イベント', '')) == '1':
        ryoukin_edit_page.reservation_banquet_event_checkbox().click()
    if str(_get_value(record, '予約 家事代行', '')) == '1':
        ryoukin_edit_page.reservation_housekeeping_checkbox().click()
    if str(_get_value(record, '実施 デイケア', '')) == '1':
        ryoukin_edit_page.implementation_daycare_checkbox().click()
    if str(_get_value(record, '実施 デイケア日数追加', '')) == '1':
        ryoukin_edit_page.implementation_daycare_days_add_checkbox().click()
    if str(_get_value(record, '実施 教室日数追加', '')) == '1':
        ryoukin_edit_page.implementation_classroom_days_add_checkbox().click()
    if str(_get_value(record, '実施 教室', '')) == '1':
        ryoukin_edit_page.implementation_classroom_checkbox().click()
    if str(_get_value(record, '実施 一時預かり/ホテル託児', '')) == '1':
        ryoukin_edit_page.implementation_temporary_care_checkbox().click()
    if str(_get_value(record, '実施 AKC', '')) == '1':
        ryoukin_edit_page.implementation_akc_checkbox().click()
    if str(_get_value(record, '実施 バックアップ', '')) == '1':
        ryoukin_edit_page.implementation_backup_checkbox().click()
    if str(_get_value(record, '実施 劇団四季', '')) == '1':
        ryoukin_edit_page.implementation_gekidan_shiki_checkbox().click()
    if str(_get_value(record, '実施 宴会イベント', '')) == '1':
        ryoukin_edit_page.implementation_banquet_event_checkbox().click()
    if str(_get_value(record, '実施 家事代行', '')) == '1':
        ryoukin_edit_page.implementation_housekeeping_checkbox().click()
    if str(_get_value(record, '割引チケット', '')) == '1':
        ryoukin_edit_page.discount_ticket_checkbox().click()

    # 顧客区分
    if str(_get_value(record, 'プレミア', '')) == '1':
        ryoukin_edit_page.customer_kbn_premium_checkbox().click()
    if str(_get_value(record, 'ビジター', '')) == '1':
        ryoukin_edit_page.customer_kbn_visitor_checkbox().click()
    if str(_get_value(record, '法人メンバー', '')) == '1':
        ryoukin_edit_page.customer_kbn_corporate_member_checkbox().click()
    if str(_get_value(record, '支店会員', '')) == '1':
        ryoukin_edit_page.customer_kbn_branch_member_checkbox().click()
    if str(_get_value(record, '法人', '')) == '1':
        ryoukin_edit_page.customer_kbn_corporate_checkbox().click()
    if str(_get_value(record, 'BS会員', '')) == '1':
        ryoukin_edit_page.customer_kbn_bs_member_checkbox().click()
    if str(_get_value(record, '優待会員', '')) == '1':
        ryoukin_edit_page.customer_kbn_special_member_checkbox().click()


def _get_value(record, key, default):
    val = record.get(key)
    if val is None:
        return default
    return val


# ===== main =======================================================================================================
args = sys.argv
file_path = args[1]
env = args[2] if len(args) > 2 else 'dev'

read_pages(env, file_path)
